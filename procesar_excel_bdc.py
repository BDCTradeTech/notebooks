import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

FAMILIAS = [
    'Intel Celeron', 'Intel Pentium', 'Intel Core 5', 'Intel Core 7', 'Intel Core i3', 'Intel Core i5', 'Intel Core i7', 'Intel Core i9',
    'Intel Core Ultra5', 'Intel Core Ultra7', 'Intel Core Ultra9', 'AMD Ryzen 3', 'AMD Ryzen 5', 'AMD Ryzen 7', 'AMD Ryzen 9', 'Apple'
]
PANTALLAS = [
    ('11.6', '11.6"'), ('14.1', '14.1"'), ('13"', '13"'), ('15.6', '15.6"'), ('16 inch', '16"'), ('13.3 inch', '13.3"'), ('13IN', '13"')
]
MEMORIAS = [
    ('8G', '8GB'), ('4GB', '4GB'), ('12GB', '12GB'), ('16GB', '16GB'), ('32GB', '32GB'), ('24GB', '24GB')
]
DISCOS = [
    ('512G', '512GB'), ('128G', '128GB'), ('64GB', '64GB'), ('256GB', '256GB'), ('127GB', '127GB'), ('1TB', '1TB')
]
HEADER = ['SKU', 'Marca', 'Descripción', 'Familia', 'Pantalla', 'Memoria', 'Disco', 'Qty', 'Precio', 'ETA', 'MOQ']

def extraer_valor(descripcion, opciones, default='-'):
    for pattern, value in opciones:
        if pattern in descripcion:
            return value
    return default

def procesar_excel(archivo_entrada, margen_bdc=5.0, return_path=False):
    wb = openpyxl.load_workbook(archivo_entrada)
    ws = wb.active

    # Buscar la fila de cabecera
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and all(x in [str(cell).lower() for cell in row if cell] for x in ['sku', 'marca', 'qty', 'price', 'eta', 'moq']):
            header_row = i
            break
    if not header_row:
        raise Exception("No se encontró la fila de cabecera.")

    # Mapear columnas
    header_cells = [str(cell).strip().lower() for cell in ws[header_row]]
    col_map = {name: idx for idx, name in enumerate(header_cells)}

    # Procesar filas
    datos = []
    familia_actual = '-'
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Detectar si es una fila de grupo (familia)
        if row and str(row[0]).strip() in FAMILIAS:
            familia_actual = str(row[0]).strip()
            continue
        # Saltar filas antes de la cabecera
        if ws.iter_rows(min_row=header_row, max_row=header_row):
            if row == tuple(ws[header_row-1][cell].value for cell in range(len(ws[header_row-1]))):
                continue
        # Procesar filas de datos
        if row and all(row[col_map.get(x, -1)] for x in ['sku', 'marca', 'qty', 'price', 'eta', 'moq']):
            descripcion = str(row[col_map.get('descripcion', col_map.get('descripción', -1))] or '')
            pantalla = extraer_valor(descripcion, PANTALLAS)
            memoria = extraer_valor(descripcion, MEMORIAS)
            disco = extraer_valor(descripcion, DISCOS)
            precio = float(row[col_map['price']]) * (1 + margen_bdc / 100)
            datos.append([
                row[col_map['sku']],
                row[col_map['marca']],
                descripcion,
                familia_actual,
                pantalla,
                memoria,
                disco,
                row[col_map['qty']],
                precio,
                row[col_map['eta']],
                row[col_map['moq']]
            ])

    # Crear nuevo archivo
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Notebooks'

    # Escribir cabecera
    ws_out.append(HEADER)

    # Escribir datos
    for fila in datos:
        ws_out.append(fila)

    # Estilos
    thin = Side(border_style='thin', color='000000')
    thick = Side(border_style='thick', color='000000')
    for row in ws_out.iter_rows():
        for cell in row:
            cell.font = Font(name='Verdana', size=11)
            cell.fill = PatternFill('solid', fgColor='FFFFFF')
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Cabecera: fondo negro, texto blanco, negrita
    for cell in ws_out[1]:
        cell.font = Font(name='Verdana', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='000000')

    # Borde exterior grueso
    max_row = ws_out.max_row
    max_col = ws_out.max_column
    for col in range(1, max_col+1):
        ws_out.cell(row=1, column=col).border = Border(top=thick, left=thin, right=thin, bottom=thin)
        ws_out.cell(row=max_row, column=col).border = Border(bottom=thick, left=thin, right=thin, top=thin)
    for row in range(1, max_row+1):
        ws_out.cell(row=row, column=1).border = Border(left=thick, top=thin, bottom=thin, right=thin)
        ws_out.cell(row=row, column=max_col).border = Border(right=thick, top=thin, bottom=thin, left=thin)
    # Esquinas
    ws_out.cell(row=1, column=1).border = Border(top=thick, left=thick, right=thin, bottom=thin)
    ws_out.cell(row=1, column=max_col).border = Border(top=thick, right=thick, left=thin, bottom=thin)
    ws_out.cell(row=max_row, column=1).border = Border(bottom=thick, left=thick, right=thin, top=thin)
    ws_out.cell(row=max_row, column=max_col).border = Border(bottom=thick, right=thick, left=thin, top=thin)

    # Formato moneda para Precio
    for row in ws_out.iter_rows(min_row=2, min_col=9, max_col=9, max_row=max_row):
        for cell in row:
            cell.number_format = '"$"#,##0.00'

    # Guardar archivo
    fecha = datetime.now().strftime('%d-%m-%y')
    nombre_archivo = f'Notebooks BDC {fecha}.xlsx'
    output_path = os.path.join(os.path.dirname(archivo_entrada), nombre_archivo)
    wb_out.save(output_path)
    print(f'Archivo generado: {output_path}')
    if return_path:
        return output_path

if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('Uso: python procesar_excel_bdc.py archivo_entrada.xlsx [margen_bdc]')
    else:
        archivo = sys.argv[1]
        margen = float(sys.argv[2]) if len(sys.argv) > 2 else 5.0
        procesar_excel(archivo, margen) 